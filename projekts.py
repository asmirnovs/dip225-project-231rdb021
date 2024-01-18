# ----- PART 1: GETTING INFO FROM EMAILS -----
import imaplib
import email
import re
from datetime import datetime

# --- logging into the email ---
gmail = imaplib.IMAP4_SSL("imap.gmail.com")
gmail.login(login, password)
gmail.select('INBOX')

# --- fetch mail ---
for i in range(3): # ask user for date
    since = input("Enter date (dd-Mon-YYYY): ")
    try: datetime.strptime(since, "%d-%b-%Y"); break # check if the format is correct
    except ValueError: print('Error, wrong date format')
    if i==2: quit() # quit if failed to provide date

value = 'info@citybee.lv'
print('Searching mail...')
_, data = gmail.search(None, 'FROM', value, f'(SINCE {str(since)})', 'BODY "par sniegtajiem pakalpojumiem"') # fetch mail
data = data[0].split()
print('Search done')

res = []

# --- get link and password from email ---
print('Fetching email...', end="")
for mail in data:
    _, msg = gmail.fetch(mail, '(RFC822)')
    msg = email.message_from_bytes(msg[0][1])
    for part in msg.walk():
        if part.get_content_type() == 'text/plain':
            body = part.get_payload(None, True).decode()
            url = re.search("Pārskatīt rēķinu (.+)", body).group().lstrip("Pārskatīt rēķinu (").replace(" )", "").strip().strip(" ")
            parole = re.search("Rēķinu var apskatīt tikai pēc paroles ievadīšanas.+\n", body).group().lstrip("Rēķinu var apskatīt tikai pēc paroles ievadīšanas")[2:].strip()
            date = msg['date'][:-12] # get mail date
            date = datetime.strftime(datetime.strptime(date, "%a, %d %b %Y %H:%M:%S"), "%d.%m.%Y") # dd.mm.YYYY
            res.append([url, parole, date]) # add info to list
    print(".", end="", flush=True)
print('\nFetch done')
gmail.close()


# ----- PART 2: GET INFO FROM WEBSITE -----
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, NoSuchElementException, NoSuchWindowException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

cars = []
total = 0.00
xl=2

driver = webdriver.Chrome(service=Service(), options=webdriver.ChromeOptions())
wb = Workbook()
ws=wb.active

# --- headers ---
ws['A1'] = 'Date'
ws['B1'] = 'Car info'
ws['C1'] = 'Cost'
for i in ['A', 'B', 'C']: ws[f'{i}1'].font = Font(bold=True)

for doc in res:
    # --- open url ---
    driver.get(str(doc[0]))
    for _ in range(3): # 3 attempts to get info
        try:
            driver.refresh()
            WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.NAME, "invoicePassword")))

            # --- enter password and open doc ---
            driver.find_element(By.NAME, "invoicePassword").send_keys(doc[1].strip())
            driver.find_element(By.CLASS_NAME, "mdc-button__label").click()
            WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".report-box.version2.mt-32.ng-star-inserted")))

            # --- get total cost ---
            subtotal = driver.find_elements(By.CSS_SELECTOR, '.text-right.emphasis.bolder')
            for i in subtotal:
                if i.get_attribute("innerHTML") == "Kopā / Total incl. VAT, EUR":
                    break
            subtotal = subtotal[subtotal.index(i)+1].get_attribute("innerHTML") # TOTAL COST
            subtotal = float(subtotal)

            # --- get list with all the cars ---
            for i in driver.find_elements(By.CSS_SELECTOR, '.invoice-line.ng-star-inserted'):
                row = i.find_elements(By.TAG_NAME, "td")
                ws['A'+str(xl)] = row[2].get_attribute("innerHTML")[:11].strip()
                ws['B'+str(xl)] = row[1].get_attribute("innerHTML")
                ws['C'+str(xl)] = str(row[3].get_attribute("innerHTML")) + '€'
                xl += 1
            # --- add info to the spreadsheet ---
            ws.merge_cells(start_row=xl, start_column=2, end_row=xl, end_column=3)
            ws['A'+str(xl)] = doc[2]
            ws['A'+str(xl)].font = Font(bold=True)
            ws['B'+str(xl)] = f"---- INVOICE SUBTOTAL: {round(subtotal, 2)}€ ----"
            ws['B'+str(xl)].font = Font(bold=True)
            xl += 1
            total += round(subtotal, 2)
            wb.save('cars.xlsx')
            break
        except TimeoutException or NoSuchElementException:
            continue # refresh and try again 
        except KeyboardInterrupt or NoSuchWindowException:
            print('Process stopped')
            quit()
# --- add total cost to the spreadsheet ---
ws.merge_cells(start_row=xl, start_column=1, end_row=xl, end_column=3)
ans = f"\n TOTAL COST: {round(total, 2)}€"
ws['A'+str(xl)] = ans
wb.save('cars.xlsx')

wb.close()
driver.close()

print(ans) # print total cost to console for convenience